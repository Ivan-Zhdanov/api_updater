import time
import cv2
import numpy as np
import pyautogui
import pandas as pd
import pyperclip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

def search_pic_x_y(image_name, tr):
    screenshot = pyautogui.screenshot()
    screenshot.save('screenshot.jpg')

    # Преобразование скриншота в массив numpy
    screenshot_np = np.array(screenshot)

    # Преобразование BGR формата (используемого OpenCV) в RGB формат
    screenshot_cv2 = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2BGR)

    # Отображение скриншота
    # cv2.imshow('Screenshot', screenshot_cv2)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()


    # Загрузка изображения
    image = cv2.imread('screenshot.jpg')
    # Преобразование изображения в оттенки серого (grayscale)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Загрузка изображения, содержащего объект, который нужно найти (шаблон)
    template = cv2.imread(image_name, 0)

    # Поиск соответствия шаблона на изображении
    result = cv2.matchTemplate(gray, template, cv2.TM_CCOEFF_NORMED)

    # Установка порога для определения соответствий
    threshold = tr
    loc = np.where(result >= threshold)
    coordinates = []
    for pt in zip(*loc[::-1]):
        center_x = pt[0] + template.shape[1] // 2
        center_y = pt[1] + template.shape[0] // 2
        coordinates.append((center_x, center_y))
    try:
        print(coordinates[0], coordinates[1])
    except:
        pass

    # # Отрисовка прямоугольников вокруг найденных соответствий
    # for pt in zip(*loc[::-1]):
    #     cv2.rectangle(image, pt, (pt[0] + template.shape[1], pt[1] + template.shape[0]), (0, 255, 0), 2)
    # print(pt)
    # pyautogui.moveTo(pt)

    try:
        pyautogui.click(center_x, center_y)
    except:
        pass
    return (center_x, center_y)


def take_api(email, passw):
    # Переключение на браузер Мозилла
    # try:
    #     search_pic_x_y('mozz.jpg', 0.8)
    #     time.sleep(2)
    # except:
    #     time.sleep(3)
    #     pyautogui.moveTo(500, 500)
    #     time.sleep(4)
    #     search_pic_x_y('mozz.jpg', 0.8)
    #     time.sleep(1)

    # ввести взятые значения из Логин в строке email браузера
    time.sleep(2)
    try:
        # search_pic_x_y('email.jpg', 0.7)
        search_pic_x_y('email_ru.jpg', 0.7)
        pyautogui.write(email)
        pyautogui.press('enter')
        pyautogui.press('enter')
    except:
        search_pic_x_y('email3.jpg', 0.7)
        pyautogui.write(email)
        pyautogui.press('enter')
        pyautogui.press('enter')


    time.sleep(5)
    pyautogui.press('enter')
    pyautogui.press('enter')
    try:
        # search_pic_x_y('cont.jpg', 0.7)
        search_pic_x_y('cont_ru.jpg', 0.75)
        pyautogui.write(passw)
        pyautogui.press('enter')
    except:
        pass
    time.sleep(3)
    try:
        search_pic_x_y('api_arrow.jpg', 0.6)
    except:
        try:
            search_pic_x_y('api_arrow3.jpg', 0.8)
        except:
            try:
                pyautogui.moveTo(x=1162, y=510)
                time.sleep(1)
                pyautogui.click()
            except:
                pass
    time.sleep(4)
    try:
        time.sleep(3)
        search_pic_x_y('api_lock.jpg', 0.7)
    except:
        try:
            time.sleep(3)
            search_pic_x_y('api_lock4.jpg', 0.7)
        except:
            time.sleep(3)
            pyautogui.moveTo(38, 337)
            time.sleep(2)
            pyautogui.click()
    time.sleep(4)
    search_pic_x_y('btn_create.jpg', 0.7)
    time.sleep(4)
    search_pic_x_y('btn_create_green.jpg', 0.7)

    time.sleep(4)
    search_pic_x_y('btn_copy.jpg', 0.7)
    time.sleep(4)
    pyautogui.press('enter')

    # Получили скопированный API-KEY
    clipboard_text = pyperclip.paste()
    print(clipboard_text)

    # Выход из аккаунта openai
    time.sleep(4)
    search_pic_x_y('personal.jpg', 0.7)
    time.sleep(2)
    search_pic_x_y('log_out.jpg', 0.7)

    time.sleep(4)
    pyautogui.doubleClick(200, 200)

    return clipboard_text

def read_excel(i):
    excel_file_path = 'OpenAI.xlsx'
    df = pd.read_excel(excel_file_path)
    email = df.iloc[i-1, df.columns.get_loc('LOGIN')]
    passw = df.iloc[i-1, df.columns.get_loc('PASSWORD')]
    print(email)
    print(passw)
    return [email, passw]


def write_excel(i, api):
    book = load_workbook('OpenAI.xlsx')
    sheet = book.active
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    sheet[f'C{i+1}'].fill = green_fill
    sheet[f'C{i+1}'].value = api
    book.save('OpenAI.xlsx')
    return 0

# --------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) <3:
        low = int(input('введите нижнюю границу'))
        hight = int(input('ведите верхнюю границу'))
    else:
        low = int(sys.argv[2])
        hight = int(sys.argv[3])
    print(f'Границы ({low} {hight}')

    # Переключение на браузер Мозилла
    try:
        search_pic_x_y('mozz.jpg', 0.8)
        time.sleep(2)
    except:
        pass
    for i in range(low, hight):
        time.sleep(2)
        print('Текущий номер строки API: ', i)
        emai_pass = read_excel(i)
        email = emai_pass[0]
        passw = emai_pass[1]
        api = take_api(email=email, passw=passw)
        write_excel(i, api)


    # while True:
    #     x = pyautogui.position()
    #     print(x)
    #     time.sleep(5)
